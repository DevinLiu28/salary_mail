#!/usr/bin/env python
# -*-coding:utf-8-*-

import os
import sys
import codecs
import time
import datetime
import smtplib
import configparser
from pathlib import Path
from email.mime.text import MIMEText
from email.header import Header
import openpyxl  # 用于处理xlsx文件


class ResourceManager:
    def __init__(self, app_name):
        """
        :param app_name: 应用程序名称，用于创建专属资源目录
        """
        self.app_name = app_name

        # 确定基础路径
        if getattr(sys, 'frozen', False):
            # 打包后模式 - 使用exe所在目录
            self.base_dir = Path(sys.executable).parent
        else:
            # 开发模式 - 使用项目根目录
            self.base_dir = Path(__file__).parent.parent

        # 创建标准子目录
        self.dirs = {
            'config': self.base_dir / 'config',
            'data': self.base_dir / 'data',
            'resources': self.base_dir / 'resources',
            'logs': self.base_dir / 'logs'
        }

        # 确保所有目录存在
        for dir_path in self.dirs.values():
            dir_path.mkdir(exist_ok=True)

    def get_path(self, file_type, filename, create_if_missing=False, default_content=None):
        """
        获取文件路径，可选创建缺失文件

        :param file_type: 文件类型（config/data/resources/logs）
        :param filename: 文件名（可包含子目录）
        :param create_if_missing: 是否自动创建缺失文件
        :param default_content: 创建文件时的默认内容（字符串或字节）
        :return: Path对象
        """
        if file_type not in self.dirs:
            raise ValueError(f"无效的文件类型: {file_type}")

        file_path = self.dirs[file_type] / filename

        if create_if_missing and not file_path.exists():
            file_path.parent.mkdir(exist_ok=True)

            if default_content is not None:
                mode = 'wb' if isinstance(default_content, bytes) else 'w'
                encoding = None if isinstance(default_content, bytes) else 'utf-8'

                with open(file_path, mode, encoding=encoding) as f:
                    f.write(default_content)

        return file_path

    def load_config(self, config_name):
        """加载配置文件"""
        config_path = self.get_path('config', f"{config_name}.ini",
                                    create_if_missing=True,
                                    default_content=f"[{config_name}]\nkey = value\n")

        config = configparser.ConfigParser()
        config.read(config_path)
        return config

    def load_text_file(self, filename):
        """加载文本文件"""
        file_path = self.get_path('data', filename)
        return file_path

    def load_excel(self, filename):
        """加载Excel文件"""
        file_path = self.get_path('data', filename)
        return file_path

    def get_resource(self, filename):
        """获取资源文件路径"""
        return self.get_path('resources', filename)


# 使用示例
if __name__ == "__main__":
    res_mgr = ResourceManager("MyApp")

    # 处理配置文件
    app_config = res_mgr.load_config("config")
    # db_config = res_mgr.load_config("db")

    # 处理数据文件
    attach = res_mgr.load_text_file("attach.txt")
    log= res_mgr.load_text_file("log.txt")
    signture= res_mgr.load_text_file("signture.txt")

    workbook = res_mgr.load_excel("工资条.xlsx")

    # 获取资源文件路径
    # avatar_path = res_mgr.get_resource("images/default_avatar.png")

    # 写入新文件
    # new_config_path = res_mgr.get_path('config', 'new_settings.ini',
    #                                    create_if_missing=True,
    #                                    default_content="[DEFAULT]\nversion=1.0")
    #
 # current_dir = os.path.dirname(os.path.abspath(__file__))


# def get_config_path(config_name):
#     # 获取可执行文件所在目录
#     if getattr(sys, 'frozen', False):
#         base_dir = Path(sys.executable).parent
#     else:
#         base_dir = Path(__file__).parent
#
#     config_path = base_dir / 'config' / f'{config_name}.ini'
#
#     # 如果不存在则创建默认配置
#     if not config_path.exists():
#         config_path.parent.mkdir(exist_ok=True)
#         with open(config_path, 'w') as f:
#             f.write(f"[{config_name}]\nkey = value\n")
#
#     return config_path

log_path = log


def loginfo(msg):
    with codecs.open(log_path, 'a', 'utf-8') as f:
        f.write(time.strftime("%Y-%m-%d %X") + "-" + msg + os.linesep)


def send_mail(to_addr, subject, html_template, user_mail, user_passwd, smtp_server, smtp_port, enable_ssl):
    try:
        message = MIMEText(html_template, 'html', 'utf-8')
        message['From'] = Header(user_mail)
        message['To'] = Header(to_addr)
        message['Subject'] = Header(subject, 'utf-8')
        # mail_obj = None
        # if enable_ssl:
        #     mail_obj = smtplib.SMTP_SSL(smtp_server)
        # else:
        mail_obj = smtplib.SMTP(smtp_server)
        mail_obj.ehlo()
        mail_obj.starttls()
        mail_obj.login(user_mail, user_passwd)
        mail_obj.sendmail(user_mail, to_addr, message.as_string())
        mail_obj.quit()
        return True
    except Exception as e:
        loginfo('发送邮箱 ' + str(to_addr) + ' 发生异常错误: ' + str(e))
        return False


def read_data(excel_file):
    excel_data = []
    # occupy lines for each item(table header,staff salary,staff salary....)
    item_lines_arr = []
    wb = openpyxl.load_workbook(filename=excel_file, read_only=False, data_only=True)
    ws = wb.worksheets[0]
    for row in ws.rows:
        row_cells = []
        for index, cell in enumerate(row):
            cell_merge = get_cell_merge(cell.row, cell.column, ws.merged_cells)
            if index == 0:
                if cell_merge["type"] == 'rowspan':
                    item_lines_arr.append(cell_merge["rowspan"])
                elif cell_merge["type"] == 'normal':
                    item_lines_arr.append(1)
            row_cells.append({
                "value": cell.value,
                "coordinate": cell.coordinate,
                "col": cell.column,
                "row": cell.row,
                "merge": cell_merge
            })
        excel_data.append(row_cells)
    return excel_data, item_lines_arr


def read_attach():
    attach_path = attach
    if os.path.exists(attach_path):
        with open(attach_path, 'r', encoding='utf-8', errors='ignore') as f:
            return f.read()
    else:
        return ''

def read_signture():
    signture_path = signture
    if os.path.exists(signture_path):
        with open(signture_path, 'r', encoding='utf-8', errors='ignore') as f:
            return f.read()
    else:
        return ''

def get_cell_merge(row, col, merged_cells):
    for item in merged_cells.ranges:
        # on the same column
        if item.min_col == item.max_col == col:
            # rowspan
            if item.min_row == row:
                return {"type": "rowspan", "rowspan": item.max_row - item.min_row + 1}
            elif item.min_row < row <= item.max_row:
                return {"type": "none"}
        # on the same row
        elif item.max_row == item.min_row == row:
            # colspan
            if item.min_col == col:
                return {"type": "colspan", "colspan": item.max_col - item.min_col + 1}
            elif item.min_col < col <= item.max_col:
                return {"type": "none"}
        elif item.min_row == row and item.min_col == col:
            return {"type": "mix", "rowspan": item.max_row - item.min_row + 1,
                    "colspan": item.max_col - item.min_col + 1}
        elif item.min_row <= row <= item.max_row and item.min_col <= col <= item.max_col:
            return {"type": "none"}
    return {"type": "normal"}


def fill_table(row_datas, style):
    grid = 'td' if style == 'td' else 'th'
    holder_str = ''
    for row_cells in row_datas:
        holder_str += '<tr>'
        for cell in row_cells[1:]:
            try:
                val = '' if cell["value"] is None else cell["value"]
            except Exception as e:
                print(e)
            if cell["merge"]["type"] == 'rowspan':
                holder_str += '<%s style="padding-left:20px;padding-right:20px;" rowspan="%s">%s</%s>'\
                              % (grid, cell["merge"]["rowspan"], val, grid)
            if cell["merge"]["type"] == 'colspan':
                holder_str += '<%s style="text-align:center;" colspan="%s">%s</%s>' \
                              % (grid, cell["merge"]["colspan"], val, grid)
            if cell["merge"]["type"] == 'mix':
                holder_str += '<%s style="text-align:center;" rowspan="%s" colspan="%s">%s</%s>'\
                              % (grid, cell["merge"]["rowspan"], cell["merge"]["colspan"], val, grid)
            if cell["merge"]["type"] == 'none':
                pass
            if cell["merge"]["type"] == 'normal':
                holder_str += '<%s style="padding-left:20px;padding-right:20px;">%s</%s>'\
                              % (grid, val, grid)
        holder_str += '</tr>'
    return holder_str


def main():
    # cf = configparser.ConfigParser()
    # cf.read(app_config)
    user = app_config.get('user', 'email')
    pwd = app_config.get('user', 'password')
    server = app_config.get('user', 'smtp_server')
    port = app_config.getint('user', 'smtp_port')
    enable_ssl = app_config.getboolean('user', 'enable_ssl')

    excel_data, item_lines_arr = read_data(workbook)
    staff_index = item_lines_arr[0]
    user_name = excel_data[staff_index][3]["value"]

    today_day = datetime.date.today()
    # today_month = datetime.datetime.now().month

    salary_time = excel_data[staff_index][1]["value"]
    print('The Company paid wages before the 10th')
    print('Today is ' + time.strftime("%B %d"))
    mail_subject = "请查收："+salary_time+"工资条"
    # Pay money before the 10th of each month
    # if today_day > 10:
    #     mail_subject = mail_subject % today_month
    # else:
    #     today_month = today_month - 1
    #     if today_month == 0:
    #         today_month = 12
    #     mail_subject = mail_subject % today_month
    # english_month = datetime.date(1900, today_month, 1).strftime('%B')
    # print('The mail subject will be show as "' + english_month + ' salley bill"')
    # print("\n")

    attach_text = read_attach()
    signture_text = read_signture()
    html_template = 'Dear， <<username_placeholder>> '
    html_template += '<pre>' + attach_text + '</pre><br/>' if attach_text else ''
    html_template += '<table border="1px solid black">'
    html_template += '<thead>'
    html_template += '<<header_placeholder>>'
    html_template += '</thead>'
    html_template += '<tbody>'
    html_template += '<<salary_placeholder>>'
    html_template += '</tbody>'
    html_template += '</table>'
    html_template += '<div><span></span><div>&nbsp;</div>' + str(today_day) + '</div>'
    html_template += '<hr color="#b5c4df" size="1" align="left" style="width: 210px; height: 1px;">'
    html_template += signture_text
    header_datas = excel_data[0:item_lines_arr[0]]
    holder_str = fill_table(header_datas, 'th')
    html_template = html_template.replace('<<header_placeholder>>', holder_str)

    has_failture = False
    for staff_lines in item_lines_arr[1:]:
        staff_email = excel_data[staff_index][0]["value"]
        staff_user = excel_data[staff_index][3]["value"]
        staff_datas = excel_data[staff_index:staff_index + staff_lines]
        holder_str = fill_table(staff_datas, 'td')
        html_content = html_template.replace('<<username_placeholder>>', staff_user).replace('<<salary_placeholder>>', holder_str)
        if staff_email is not None:
            staff_email = staff_email.replace("\n", "").replace("\r", "").replace(" ", "")
            send_result = send_mail(staff_email, mail_subject, html_content, user, pwd, server, port, enable_ssl)
            if not send_result:
                has_failture = True
                loginfo('员工：' + str(staff_user) + ' 邮箱:' + str(staff_email) + ' 发送失败!!!,请重新发送！')
                print('员工：' + str(staff_user) + ' 邮箱:' + str(staff_email) + ' 发送失败!!!,请重新发送！')
            else:
                loginfo('员工：' + str(staff_user) + ' 邮箱:' + str(staff_email) + ' 发送成功！')
                print('员工：' + str(staff_user) + ' 邮箱:' + str(staff_email) + ' 发送成功！')
                time.sleep(1)
        staff_index += staff_lines
    print("\n")
    if has_failture:
        print("有部分邮件发送失败, 请在日志文件log.txt中检查")
        print("\n")
        input('按任意键退出...')
    else:
        print("程序运行成功，所有邮件已发送.")
        print('程序三秒将退出...')
        time.sleep(3)
    sys.exit(0)


main()
